# -*- coding: utf-8 -*-
"""
src/medios_magneticos.py
========================
Generación de Información Exógena (Medios Magnéticos) para la DIAN
de Colombia.

Formatos soportados:
  - 1001: Pagos o abonos en cuenta y retenciones practicadas
  - 1003: Retenciones que nos practicaron (en ventas)
  - 1004: Descuentos tributarios
  - 1005: IVA descontable
  - 1006: IVA generado
  - 1007: Ingresos recibidos
  - 1008: Cuentas por cobrar al 31 de diciembre
  - 1009: Cuentas por pagar al 31 de diciembre
  - 1010: Información de socios y accionistas
  - 1011: Información declaraciones tributarias (consolidado)
  - 1012: Información de saldos (caja, bancos, inversiones)

NOTAS:
  - Los formatos 1010, 1011 y 1012 requieren información que NO siempre
    está completa en Odoo. Se generan plantillas con los datos disponibles
    y el usuario debe completar manualmente lo que falte.
  - El mapeo PUC → concepto DIAN es GENÉRICO. Cada empresa puede
    requerir ajustes según su catálogo contable específico.
"""
from __future__ import annotations

import re
from datetime import date
from typing import Optional

import pandas as pd

from .financial_statements import enrich_chart_with_puc


# Patrones regex para extraer cédula/NIT desde texto libre (ref del asiento,
# nombre de línea, etc.). Importante para casos donde el partner no tiene
# documento registrado en su ficha pero la cédula aparece en la referencia.
#
# Ejemplos que matchean:
#   "Nomina de 1077468686-01"         → 1077468686
#   "Pago nómina CC 1077468686"       → 1077468686
#   "Factura proveedor NIT 900123456" → 900123456
#   "C.C. 52123456"                   → 52123456
_PATRONES_DOC = [
    re.compile(r"(?:nomina\s+de\s+|nómina\s+de\s+)(\d{6,15})", re.IGNORECASE),
    re.compile(r"(?:^|\W)(?:c\.?\s*c\.?|cc|cedula|cédula)[\s:.\-]*(\d{6,15})", re.IGNORECASE),
    re.compile(r"(?:^|\W)(?:nit|n\.?i\.?t\.?)[\s:.\-]*(\d{6,15})", re.IGNORECASE),
    re.compile(r"(?:^|\W)(?:c\.?\s*e\.?|ce)[\s:.\-]*(\d{6,15})", re.IGNORECASE),
]


def _extract_doc_from_text(text: str) -> str:
    """
    Intenta extraer una cédula/NIT desde texto libre (referencia, descripción).
    Devuelve el número o cadena vacía si no encuentra.
    """
    if not text or not isinstance(text, str):
        return ""
    s = text.strip()
    if not s or s.lower() == "false":
        return ""
    for pat in _PATRONES_DOC:
        m = pat.search(s)
        if m:
            doc = m.group(1)
            # Validar longitud razonable (cédula CO: 6-10 dígitos, NIT: 9-10)
            if 6 <= len(doc) <= 15:
                return doc
    return ""


def _inferir_documentos_desde_refs(moves: pd.DataFrame) -> dict[int, str]:
    """
    Construye un dict {partner_id: cedula_inferida} a partir de los campos
    `ref` y `name` (descripción de la línea) de los movimientos.

    Para cada partner_id, busca el primer texto donde aparezca un patrón
    de documento válido. Si encuentra varios distintos, usa el más
    frecuente (probabilidad de ser el correcto).
    """
    if moves is None or moves.empty or "partner_id" not in moves.columns:
        return {}

    # Construir texto combinado por partner_id desde varios campos
    text_cols = [c for c in ["ref", "name", "move_id_name"] if c in moves.columns]
    if not text_cols:
        return {}

    # Para cada partner_id, recolectar candidatos
    candidates: dict[int, list[str]] = {}
    sub = moves[moves["partner_id"].notna()].copy()
    for _, row in sub.iterrows():
        try:
            pid = int(row["partner_id"])
        except (TypeError, ValueError):
            continue
        for col in text_cols:
            val = row.get(col)
            doc = _extract_doc_from_text(str(val) if val else "")
            if doc:
                candidates.setdefault(pid, []).append(doc)
                break  # un doc por línea, no replicar

    # Para cada partner, escoger el documento más frecuente
    result: dict[int, str] = {}
    for pid, docs in candidates.items():
        if not docs:
            continue
        # Conteo de frecuencias
        counter: dict[str, int] = {}
        for d in docs:
            counter[d] = counter.get(d, 0) + 1
        # Tomar el más frecuente
        best = max(counter.items(), key=lambda kv: kv[1])
        result[pid] = best[0]
    return result


def _enrich_moves_with_puc(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
) -> pd.DataFrame:
    """
    Enriquece movimientos contables con columnas de clasificación PUC.

    GARANTÍA: filtra explícitamente parent_state == 'posted' para excluir
    asientos en borrador o cancelados, aunque el extractor ya lo haga
    server-side (doble validación defensiva).

    Agrega:
      - account_code: código real (o inferido desde account_id_name)
      - account_name: nombre de la cuenta
      - puc_group: '1', '2', '3', '4', '5', '6', '7'
      - puc_subgroup: '11', '14', '21', '22', '23', '24', '41', '51', etc.

    Usa la misma lógica robusta que financial_statements: si el `code`
    de la cuenta no es código PUC, infiere desde `account_type` y nombre.
    """
    if moves is None or moves.empty:
        return moves

    # DOBLE VALIDACIÓN: filtrar parent_state=posted si la columna existe.
    # El extractor ya lo hace server-side pero esto garantiza que si en
    # el futuro alguien cambia el extractor, los formatos DIAN sigan
    # excluyendo borradores y cancelados.
    moves = moves.copy()
    if "parent_state" in moves.columns:
        moves = moves[moves["parent_state"] == "posted"]
    elif "state" in moves.columns:
        moves = moves[moves["state"] == "posted"]
    if chart is None or chart.empty:
        df = moves.copy()
        df["account_code"] = df.get("account_code", "").astype(str)
        df["puc_group"] = ""
        df["puc_subgroup"] = ""
        return df

    # Enriquecer chart con clasificación PUC
    chart_e = enrich_chart_with_puc(chart)
    keep_cols = [c for c in [
        "id", "code", "name", "puc_group", "puc_subgroup",
        "subgrupo", "account_type",
    ] if c in chart_e.columns]
    chart_min = chart_e[keep_cols].rename(columns={
        "id": "account_id",
        "code": "account_code",
        "name": "account_name",
    })

    df = moves.copy()
    # Si moves ya tiene account_code (de un merge previo), no lo pisamos
    if "account_code" in df.columns and df["account_code"].notna().any():
        # Hacer merge solo para añadir puc_group/puc_subgroup
        cols_to_add = [
            "account_id", "puc_group", "puc_subgroup",
        ]
        if "account_name" not in df.columns:
            cols_to_add.append("account_name")
        df = df.merge(chart_min[cols_to_add], on="account_id", how="left")
    else:
        df = df.merge(chart_min, on="account_id", how="left")

    # Defaults seguros — Odoo devuelve False (booleano) para valores
    # nulos vía XML-RPC. fillna+astype convertiría a "False" literal,
    # así que primero reemplazamos False por None y luego rellenamos.
    for col in ("account_code", "account_name", "puc_group", "puc_subgroup"):
        if col in df.columns:
            df[col] = df[col].replace({False: None, "False": None})
            df[col] = df[col].fillna("").astype(str)

    # FALLBACK: si account_code quedó vacío pero account_id_name SÍ existe,
    # extraer el código del display_name de la cuenta. Probar múltiples
    # formatos que Odoo puede usar:
    #   "510515 Sueldos"      → 510515
    #   "510515 - Sueldos"    → 510515
    #   "510515-Sueldos"      → 510515
    #   "[510515] Sueldos"    → 510515
    #   "Sueldos (510515)"    → 510515
    #   "Sueldos"             → "" (no recuperable)
    if "account_code" in df.columns and "account_id_name" in df.columns:
        codigo_vacio = df["account_code"].astype(str).isin(["", "False", "nan", "None"])
        if codigo_vacio.any():
            patrones_codigo = [
                re.compile(r"^\s*(\d{4,15})\b"),           # "510515 Sueldos"
                re.compile(r"^\s*\[\s*(\d{4,15})\s*\]"),   # "[510515] Sueldos"
                re.compile(r"\((\d{4,15})\)\s*$"),          # "Sueldos (510515)"
                re.compile(r"\b(\d{4,15})\s*[-:]"),         # "510515-Sueldos"
            ]

            def _extraer_codigo(nombre):
                if not nombre or str(nombre).lower() in ("false", "nan", "none"):
                    return ""
                s = str(nombre).strip()
                for pat in patrones_codigo:
                    m = pat.search(s)
                    if m:
                        return m.group(1)
                return ""

            df.loc[codigo_vacio, "account_code"] = df.loc[
                codigo_vacio, "account_id_name"
            ].apply(_extraer_codigo)

        # SEGUNDO FALLBACK: usar account_name si todavía está vacío
        codigo_vacio2 = df["account_code"].astype(str).isin(["", "False", "nan", "None"])
        if codigo_vacio2.any() and "account_name" in df.columns:
            patrones_codigo = [
                re.compile(r"^\s*(\d{4,15})\b"),
                re.compile(r"^\s*\[\s*(\d{4,15})\s*\]"),
                re.compile(r"\((\d{4,15})\)\s*$"),
            ]

            def _extraer_codigo2(nombre):
                if not nombre or str(nombre).lower() in ("false", "nan", "none"):
                    return ""
                s = str(nombre).strip()
                for pat in patrones_codigo:
                    m = pat.search(s)
                    if m:
                        return m.group(1)
                return ""

            df.loc[codigo_vacio2, "account_code"] = df.loc[
                codigo_vacio2, "account_name"
            ].apply(_extraer_codigo2)

    # FALLBACK 2: si account_name quedó vacío, usar account_id_name limpio
    if "account_name" in df.columns and "account_id_name" in df.columns:
        nombre_vacio = df["account_name"].astype(str).isin(["", "False", "nan", "None"])
        if nombre_vacio.any():
            def _limpiar_nombre(nombre):
                if not nombre or str(nombre).lower() in ("false", "nan", "none"):
                    return ""
                s = str(nombre).strip()
                # Quitar el código inicial ("510515 Sueldos" → "Sueldos")
                return re.sub(r"^\d{4,15}\s*[-:]?\s*", "", s).strip()
            df.loc[nombre_vacio, "account_name"] = df.loc[
                nombre_vacio, "account_id_name"
            ].apply(_limpiar_nombre)

    return df


# ===========================================================================
# Mapeo genérico PUC colombiano → Concepto DIAN
# ===========================================================================
#
# Para FORMATO 1001 (pagos a terceros), los conceptos DIAN principales son:
#   5001 Salarios, prestaciones sociales y demás pagos laborales
#   5002 Honorarios
#   5003 Comisiones
#   5004 Servicios
#   5005 Arrendamientos
#   5006 Intereses y rendimientos financieros
#   5007 Otros costos y deducciones
#   5009 Compra de bienes raíces
#   5010 Compras de activos fijos diferentes a bienes raíces
#   5011 Aportes a sistemas de seguridad social en salud
#   5012 Aportes a pensiones obligatorias
#   5016 Demás costos y deducciones (gastos diversos deducibles)
#   ... (lista completa en resoluciones DIAN anuales)
#
# Mapeo simplificado por prefijo de cuenta PUC:
CONCEPTOS_1001_POR_PUC: dict[str, str] = {
    # Gastos personales / laborales
    "5105": "5001",  # Gastos de personal → Salarios
    "5106": "5001",
    "5108": "5001",
    "7105": "5001",  # Costos de personal (CMV)
    # Honorarios
    "5110": "5002",  # Honorarios → Honorarios
    # Comisiones
    "5115": "5003",
    # Servicios técnicos / asistencia técnica / servicios generales
    "5135": "5004",  # Servicios → Servicios
    # Concepto 5016 — Demás costos y deducciones (gastos diversos)
    "5130": "5016",  # Seguros (administración)
    "5140": "5016",  # Gastos legales (administración)
    "5145": "5016",  # Mantenimiento y reparaciones (administración)
    "5150": "5016",  # Adecuación e instalación (administración)
    "5155": "5016",  # Gastos de viaje (administración)
    "5160": "5016",  # Depreciaciones (administración)
    "5165": "5016",  # Amortizaciones (administración)
    "5195": "5016",  # Diversos (administración)
    "5230": "5016",  # Seguros (ventas)
    "5240": "5016",  # Gastos legales (ventas)
    "5245": "5016",  # Mantenimiento y reparaciones (ventas)
    "5250": "5016",  # Adecuación e instalación (ventas)
    "5255": "5016",  # Gastos de viaje (ventas)
    "5260": "5016",  # Depreciaciones (ventas)
    "5265": "5016",  # Amortizaciones (ventas)
    "5295": "5016",  # Diversos (ventas)
    # Arrendamientos
    "5120": "5005",
    "5125": "5005",
    # Intereses y financieros
    "5305": "5006",
    "5310": "5006",
    "5315": "5006",
    # Compras de mercancía (CMV)
    "6135": "5009",  # Costo de ventas → Compras (genérico)
    "1435": "5009",  # Inventario → Compras
    # Activos fijos
    "1504": "5010",  # Terrenos
    "1516": "5010",  # Construcciones
    "1520": "5010",  # Maquinaria y equipo
    "1524": "5010",  # Muebles y enseres
    "1528": "5010",  # Equipo computación
    "1540": "5010",  # Vehículos
    # Default
    "_default": "5007",  # Otros costos y deducciones
}

# Para FORMATO 1007 (ingresos):
#   4001 Operacionales — venta neta de bienes/servicios
#   4002 Otros ingresos
CONCEPTOS_1007_POR_PUC: dict[str, str] = {
    "4135": "4001",  # Comercio al por mayor → Operacional
    "4140": "4001",
    "4145": "4001",
    "4150": "4001",
    "4155": "4001",
    "4160": "4001",
    "4165": "4001",
    "4170": "4002",  # Otros ingresos
    "_default": "4001",
}

# Para FORMATO 1009 (cuentas por pagar):
#   2201 Proveedores nacionales
#   2202 Proveedores del exterior
#   2335 Costos y gastos por pagar
#   2370 Retenciones y aportes laborales por pagar
#   ...
# Concepto único típicamente: 1303-1399
CONCEPTOS_1009_POR_PUC: dict[str, str] = {
    "2205": "1301",  # Proveedores → Saldo a favor proveedores
    "2335": "1306",  # Costos y gastos por pagar
    "2380": "1304",  # Acreedores varios
    "_default": "1306",
}


def _get_concepto(account_code: str, mapping: dict) -> str:
    """Mapea código PUC a concepto DIAN usando el mapeo dado."""
    if not account_code:
        return mapping.get("_default", "")
    s = str(account_code).strip()
    # Buscar primero el prefijo exacto de 4 dígitos
    if len(s) >= 4 and s[:4] in mapping:
        return mapping[s[:4]]
    # Después prefijo de 2 dígitos (grupo)
    if len(s) >= 2 and s[:2] in mapping:
        return mapping[s[:2]]
    return mapping.get("_default", "")


# ===========================================================================
# Helpers de identificación de terceros
# ===========================================================================

# Mapeo de tipos de documento DIAN
TIPO_DOC_DIAN: dict[str, str] = {
    "11": "Registro civil",
    "12": "Tarjeta de identidad",
    "13": "Cédula de ciudadanía",
    "21": "Tarjeta de extranjería",
    "22": "Cédula de extranjería",
    "31": "NIT",
    "41": "Pasaporte",
    "42": "Documento extranjero",
    "43": "Sin identificación del exterior",
}


def _get_partner_document(partner: dict) -> str:
    """
    Obtiene el documento del partner probando múltiples campos en orden
    de prioridad (importante para Odoo con localización Colombia):
      1. vat (estándar Odoo)
      2. l10n_co_document_number (localización CO)
      3. identification_document (otro módulo CO)
      4. ref (campo de referencia interna, algunos lo usan para cédula)
    """
    for field in ("vat", "l10n_co_document_number",
                  "identification_document", "ref"):
        value = partner.get(field)
        if value and str(value).strip() and str(value).strip().lower() != "false":
            return str(value).strip()
    return ""


def _infer_tipo_doc(partner: dict) -> str:
    """
    Infiere el tipo de documento DIAN desde info del partner.

    Si hay `l10n_latam_identification_type_id` (localización CO), usa
    ese mapping directamente. Si no, heurística por is_company/longitud.
    """
    # Intentar leer el tipo de localización CO
    id_type = partner.get("l10n_latam_identification_type_id")
    if isinstance(id_type, (list, tuple)) and len(id_type) > 1:
        name = str(id_type[1]).lower()
        if "nit" in name:
            return "31"
        if "cedula de ciudadan" in name or "cédula de ciudadan" in name:
            return "13"
        if "tarjeta de identidad" in name:
            return "12"
        if "cedula de extranjer" in name or "cédula de extranjer" in name:
            return "22"
        if "pasaporte" in name:
            return "41"

    is_company = partner.get("is_company", False)
    doc = _get_partner_document(partner)
    if is_company:
        return "31"  # NIT
    if doc:
        return "13"
    return "13"  # Default


def _clean_vat(vat_or_doc: str, partner_dv: Optional[str] = None) -> tuple[str, str]:
    """
    Limpia un NIT/cédula. Devuelve (número_sin_dv, dv).
    Si tiene formato '900123456-7' separa el DV.
    Si se pasa `partner_dv` separadamente (de l10n_co_dv), lo usa.
    """
    if not vat_or_doc:
        return "", ""
    v = str(vat_or_doc).replace(".", "").replace(" ", "").strip()
    if v.lower() == "false":
        return "", ""
    # Si ya viene con guión (NIT colombiano)
    if "-" in v:
        parts = v.split("-")
        return parts[0], parts[1] if len(parts) > 1 else (partner_dv or "")
    # Si no tiene guión pero tenemos DV externo, usarlo
    return v, str(partner_dv or "").strip()


def _split_nombre(partner_name: str) -> tuple[str, str, str, str]:
    """
    Divide un nombre en (apellido1, apellido2, nombre1, nombre2).
    Para personas naturales.
    """
    if not partner_name:
        return "", "", "", ""
    parts = partner_name.strip().split()
    if len(parts) == 1:
        return parts[0], "", "", ""
    if len(parts) == 2:
        return parts[0], "", parts[1], ""
    if len(parts) == 3:
        return parts[0], parts[1], parts[2], ""
    if len(parts) >= 4:
        return parts[0], parts[1], parts[2], " ".join(parts[3:])
    return "", "", "", ""


def _row_tercero(partner: dict) -> dict:
    """
    Construye los campos comunes de tercero para los formatos DIAN.

    Devuelve dict con:
      tipo_doc, numero_doc, dv,
      apellido1, apellido2, nombre1, nombre2, razon_social,
      direccion, departamento, municipio, pais
    """
    tipo_doc = _infer_tipo_doc(partner)
    doc = _get_partner_document(partner)
    dv_externo = partner.get("l10n_co_dv") or ""
    numero_doc, dv = _clean_vat(doc, partner_dv=str(dv_externo))
    name = partner.get("name", "")
    is_company = partner.get("is_company", False)

    if is_company:
        razon_social = name
        ap1 = ap2 = no1 = no2 = ""
    else:
        razon_social = ""
        ap1, ap2, no1, no2 = _split_nombre(name)

    return {
        "tipo_doc": tipo_doc,
        "numero_doc": numero_doc,
        "dv": dv,
        "apellido1": ap1,
        "apellido2": ap2,
        "nombre1": no1,
        "nombre2": no2,
        "razon_social": razon_social,
        "direccion": partner.get("street", "") or "",
        "departamento": partner.get("state_name", "") or "",
        "municipio": partner.get("city", "") or "",
        "pais": partner.get("country_name", "Colombia") or "Colombia",
    }


# ===========================================================================
# FORMATO 1001 — Pagos o abonos en cuenta y retenciones practicadas
# ===========================================================================


def build_formato_1001(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    partners: pd.DataFrame,
    year: int,
    umbral_minimo: float = 0,
) -> pd.DataFrame:
    """
    Construye Formato 1001: pagos a terceros por concepto.

    Toma movimientos de cuentas grupo 5 (gastos) y 6 (costos), agrupados
    por (tercero, concepto). El "pago" es el débito a esas cuentas.

    Retenciones: suma de movimientos a cuentas 2365 (Retención en la
    fuente) y 2367 (IVA retenido) del mismo período por cada tercero.

    Args:
        umbral_minimo: monto mínimo para incluir un pago. Por defecto 0
            (sin filtro). Ajusta según resolución DIAN.
    """
    if moves is None or moves.empty:
        return pd.DataFrame()

    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.year == year]
    if df.empty:
        return pd.DataFrame()

    # Enriquecer con clasificación PUC robusta (igual que Estados Financieros)
    df = _enrich_moves_with_puc(df, chart)

    # CRÍTICO: el Formato 1001 reporta pagos A PROVEEDORES (compras).
    # NO debe incluir líneas de facturas de venta (out_invoice/out_refund)
    # aunque toquen cuentas 5xxx/6xxx (ej. costo de ventas asociado a una
    # FEV con partner_id del cliente).
    if "move_type" in df.columns:
        # Tipos que SÍ van al 1001: compras, NC de proveedor, asientos manuales
        TIPOS_VALIDOS_1001 = {
            "in_invoice", "in_refund",  # facturas de proveedor
            "entry",                      # asientos manuales (nómina, ajustes)
            False, None, "",             # sin tipo (asientos de diario sin doc)
        }
        df = df[df["move_type"].isin(TIPOS_VALIDOS_1001) | df["move_type"].isna()]

    # Inferir documentos desde refs/descripciones para partners sin doc
    # en su ficha (ej: empleados con cédula en "Nomina de XXXXXXXXXX-01")
    docs_inferidos = _inferir_documentos_desde_refs(df)

    # Filtrar pagos a terceros: usa puc_group (más robusto que el code)
    pagos_mask = df["puc_group"].isin(["5", "6"]) | (
        df["account_code"].str.startswith("143")
    )
    pagos = df[pagos_mask & (df["partner_id"].notna())].copy()
    if pagos.empty:
        return pd.DataFrame()

    pagos["monto"] = pagos["debit"].fillna(0) - pagos["credit"].fillna(0)
    pagos["concepto"] = pagos["account_code"].apply(
        lambda c: _get_concepto(c, CONCEPTOS_1001_POR_PUC)
    )

    # ── Retenciones cruzadas por ASIENTO (move_id) ──
    # Las retenciones en Colombia siempre se contabilizan en el MISMO
    # asiento contable que el pago/factura. Por eso es más confiable
    # cruzar por move_id que por partner_id (que puede no estar bien
    # asignado en algunas líneas de retención).
    ret_mask = df["account_code"].str.startswith(("2365", "2367", "2368"))
    rets = df[ret_mask].copy()
    rets_by_move = pd.DataFrame()
    if not rets.empty and "move_id" in rets.columns:
        rets["monto_ret"] = rets["credit"].fillna(0) - rets["debit"].fillna(0)
        rets["tipo_ret"] = rets["account_code"].str[:4].map({
            "2365": "ret_fuente_renta",
            "2367": "ret_iva",
            "2368": "ret_ica",
        })
        # Pivot por move_id × tipo_ret
        rets_by_move = rets.pivot_table(
            index="move_id", columns="tipo_ret",
            values="monto_ret", aggfunc="sum", fill_value=0,
        ).reset_index()

    # Asociar retenciones al pago vía move_id, agrupado por (partner, concepto)
    if not rets_by_move.empty and "move_id" in pagos.columns:
        pagos_con_ret = pagos.merge(rets_by_move, on="move_id", how="left")
    else:
        pagos_con_ret = pagos.copy()
    for col in ("ret_fuente_renta", "ret_iva", "ret_ica"):
        if col not in pagos_con_ret.columns:
            pagos_con_ret[col] = 0
        pagos_con_ret[col] = pagos_con_ret[col].fillna(0)

    # Agrupar pagos + retenciones por (tercero, concepto).
    # Agregamos también el listado de cuentas contables usadas para
    # validación cruzada (qué códigos PUC componen ese pago).
    def _cuentas_unicas(s):
        vals = sorted({str(v).strip() for v in s.dropna() if str(v).strip()})
        return ", ".join(vals)

    grp = pagos_con_ret.groupby(
        ["partner_id", "concepto"], as_index=False,
    ).agg(
        pago_deducible=("monto", "sum"),
        ret_fuente_renta=("ret_fuente_renta", "sum"),
        ret_iva=("ret_iva", "sum"),
        ret_ica=("ret_ica", "sum"),
        cuentas_contables=("account_code", _cuentas_unicas),
        nombres_cuentas=("account_name", _cuentas_unicas) if "account_name" in pagos_con_ret.columns else ("account_code", lambda s: ""),
        n_lineas=("monto", "size"),
    )

    # Join con info de tercero
    if partners is not None and not partners.empty:
        partner_info = partners.set_index("id")
        out_rows = []
        for _, r in grp.iterrows():
            pid = r["partner_id"]
            try:
                p = partner_info.loc[int(pid)].to_dict() if pid in partner_info.index else {}
            except Exception:  # noqa: BLE001
                p = {}
            t = _row_tercero(p)
            # Si el partner NO tiene documento en su ficha pero sí lo
            # inferimos desde la referencia del asiento (ej. nóminas),
            # usar la cédula inferida.
            if not t.get("numero_doc"):
                doc_inferido = docs_inferidos.get(int(pid), "")
                if doc_inferido:
                    t["numero_doc"] = doc_inferido
                    # Si no era empresa, asumimos cédula
                    if not p.get("is_company", False):
                        t["tipo_doc"] = "13"
            row = {
                **t,
                "partner_id": int(pid) if pid is not None else None,
                "concepto": r["concepto"],
                "pago_deducible": round(float(r["pago_deducible"]), 0),
                "pago_no_deducible": 0,
                "ret_fuente_renta": round(float(r["ret_fuente_renta"]), 0),
                "ret_iva": round(float(r["ret_iva"]), 0),
                "ret_ica": round(float(r["ret_ica"]), 0),
                # Columnas de validación: cuentas contables que componen el pago
                "cuentas_contables": r.get("cuentas_contables", ""),
                "nombres_cuentas": r.get("nombres_cuentas", ""),
                "n_lineas": int(r.get("n_lineas", 0)),
            }
            out_rows.append(row)
        out = pd.DataFrame(out_rows)
    else:
        out = grp

    # Aplicar umbral si > 0
    if umbral_minimo > 0 and "pago_deducible" in out.columns:
        out = out[
            out["pago_deducible"].abs() >= umbral_minimo
        ].reset_index(drop=True)
    return out


def diagnosticar_formato_1001(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    partners: pd.DataFrame,
    year: int,
) -> dict:
    """
    Diagnóstico del Formato 1001. Identifica problemas:

    1. Retenciones SIN pago asociado en el mismo asiento (cuenta 2365/67/68
       con tercero pero sin gasto en el mismo move_id).
    2. Retenciones con partner_id distinto al del pago (puede indicar
       error en la contabilización).
    3. Terceros sin NIT o datos incompletos.

    Devuelve dict con:
      - sin_nit: DF con terceros sin NIT que tienen pagos
      - ret_huerfanas: retenciones sin pago en mismo asiento
      - ret_diferente_partner: retenciones con partner distinto al pago
    """
    if moves is None or moves.empty:
        return {}
    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.year == year]
    if df.empty:
        return {}
    df = _enrich_moves_with_puc(df, chart)

    # Excluir facturas de venta del diagnóstico 1001 (consistente con
    # el build_formato_1001)
    if "move_type" in df.columns:
        TIPOS_VALIDOS_1001 = {
            "in_invoice", "in_refund", "entry", False, None, "",
        }
        df = df[df["move_type"].isin(TIPOS_VALIDOS_1001) | df["move_type"].isna()]

    # Inferir cédulas desde refs (nóminas y similares)
    docs_inferidos = _inferir_documentos_desde_refs(df)

    # 1) Pagos a terceros
    pagos = df[
        df["puc_group"].isin(["5", "6"]) & df["partner_id"].notna()
    ].copy()
    pagos["monto"] = pagos["debit"].fillna(0) - pagos["credit"].fillna(0)

    # 2) Retenciones
    rets = df[
        df["account_code"].str.startswith(("2365", "2367", "2368"))
    ].copy()
    rets["monto_ret"] = rets["credit"].fillna(0) - rets["debit"].fillna(0)

    diag = {}

    # === A) Terceros con pagos pero SIN NIT ===
    # Devolvemos DOS DataFrames:
    #   - sin_nit (resumen): partner_id, nombre, total pagos
    #   - sin_nit_detalle: una fila por asiento contable, con
    #     move_id_name, ref, fecha, account_code, account_name, monto
    if partners is not None and not partners.empty and not pagos.empty:
        partners_idx = partners.set_index("id")
        # Identificar partners sin NIT
        partners_sin_nit_ids = set()
        partners_info = {}
        for pid in pagos["partner_id"].dropna().unique():
            try:
                p = partners_idx.loc[int(pid)].to_dict() if pid in partners_idx.index else {}
                encontrado = pid in partners_idx.index
            except Exception:  # noqa: BLE001
                p = {}
                encontrado = False
            # Usar _get_partner_document que revisa varios campos (vat,
            # l10n_co_document_number, identification_document, ref)
            doc = _get_partner_document(p)
            doc_inferido = docs_inferidos.get(int(pid), "")
            # Si tiene doc en ficha O lo inferimos de la referencia, OK
            doc_efectivo = doc or doc_inferido
            partners_info[int(pid)] = {
                "name": p.get("name", "(sin nombre)"),
                "documento": doc_efectivo,
                "doc_inferido_ref": doc_inferido,
                "vat": (p.get("vat") or "").strip() if str(p.get("vat", "")).lower() != "false" else "",
                "l10n_co_doc": (p.get("l10n_co_document_number") or "") if str(p.get("l10n_co_document_number", "")).lower() != "false" else "",
                "ref": (p.get("ref") or "") if str(p.get("ref", "")).lower() != "false" else "",
                "ident_doc": (p.get("identification_document") or "") if str(p.get("identification_document", "")).lower() != "false" else "",
                "encontrado_en_db": encontrado,
                "is_company": p.get("is_company", False),
                "active": p.get("active", True),
            }
            if not doc_efectivo:
                partners_sin_nit_ids.add(int(pid))

        # A.1) Resumen: total pagos por partner sin NIT
        pagos_sin_nit = pagos[pagos["partner_id"].isin(partners_sin_nit_ids)]
        if not pagos_sin_nit.empty:
            resumen = pagos_sin_nit.groupby("partner_id", as_index=False).agg(
                monto_pagos=("monto", "sum"),
                n_asientos=("move_id", "nunique") if "move_id" in pagos_sin_nit.columns else ("monto", "count"),
            )
            # Agregar columnas de diagnóstico para ver qué tiene el partner
            for col_key, dict_key in [
                ("nombre", "name"),
                ("vat", "vat"),
                ("l10n_co_doc", "l10n_co_doc"),
                ("ref", "ref"),
                ("ident_doc", "ident_doc"),
                ("encontrado_en_db", "encontrado_en_db"),
                ("is_company", "is_company"),
                ("active", "active"),
            ]:
                resumen[col_key] = resumen["partner_id"].map(
                    lambda p, k=dict_key: partners_info.get(int(p), {}).get(k, "")
                )
            cols_order = [
                "partner_id", "nombre", "monto_pagos", "n_asientos",
                "vat", "l10n_co_doc", "ref", "ident_doc",
                "encontrado_en_db", "is_company", "active",
            ]
            diag["sin_nit"] = resumen[
                [c for c in cols_order if c in resumen.columns]
            ].sort_values("monto_pagos", ascending=False).reset_index(drop=True)

            # A.2) Detalle de cada asiento donde aparecen
            detalle_cols = [c for c in [
                "partner_id", "date", "move_id", "move_id_name",
                "ref", "name",
                "account_code", "account_name",
                "debit", "credit", "monto",
            ] if c in pagos_sin_nit.columns]
            det = pagos_sin_nit[detalle_cols].copy()
            det["nombre_tercero"] = det["partner_id"].map(
                lambda p: partners_info.get(int(p), {}).get("name", "")
            )
            # Reordenar columnas
            cols_order = ["partner_id", "nombre_tercero", "date"]
            if "move_id_name" in det.columns:
                cols_order.append("move_id_name")
            elif "move_id" in det.columns:
                cols_order.append("move_id")
            if "ref" in det.columns:
                cols_order.append("ref")
            if "name" in det.columns:
                cols_order.append("name")
            if "account_code" in det.columns:
                cols_order.append("account_code")
            if "account_name" in det.columns:
                cols_order.append("account_name")
            cols_order.append("monto")
            det = det[[c for c in cols_order if c in det.columns]]
            det = det.sort_values(["partner_id", "date"]).reset_index(drop=True)
            diag["sin_nit_detalle"] = det
        else:
            diag["sin_nit"] = pd.DataFrame()
            diag["sin_nit_detalle"] = pd.DataFrame()

    # === B) Retenciones HUÉRFANAS: con tercero pero sin pago en el mismo asiento ===
    if not rets.empty and "move_id" in rets.columns:
        moves_con_pago = set(pagos["move_id"].dropna().unique())
        rets_con_partner = rets[rets["partner_id"].notna()].copy()
        rets_huerfanas = rets_con_partner[
            ~rets_con_partner["move_id"].isin(moves_con_pago)
        ].copy()
        if not rets_huerfanas.empty:
            grp_h = rets_huerfanas.groupby(
                "partner_id", as_index=False,
            )["monto_ret"].sum()
            if partners is not None and not partners.empty:
                grp_h["nombre"] = grp_h["partner_id"].map(
                    partners.set_index("id")["name"]
                )
            diag["ret_huerfanas"] = grp_h.sort_values(
                "monto_ret", ascending=False,
            )

    # === C) Retenciones con partner_id DIFERENTE al partner del pago ===
    if not rets.empty and not pagos.empty and "move_id" in rets.columns:
        # Partner por move (del pago)
        partner_pago_por_move = (
            pagos.dropna(subset=["partner_id"])
            .groupby("move_id")["partner_id"]
            .first()
        )
        # Partner de la retención
        rets_con_pago = rets.dropna(subset=["partner_id"]).merge(
            partner_pago_por_move.rename("partner_pago"),
            on="move_id", how="inner",
        )
        diff = rets_con_pago[
            rets_con_pago["partner_id"] != rets_con_pago["partner_pago"]
        ]
        if not diff.empty:
            if partners is not None and not partners.empty:
                partner_names = partners.set_index("id")["name"]
                diff = diff.copy()
                diff["nombre_ret"] = diff["partner_id"].map(partner_names)
                diff["nombre_pago"] = diff["partner_pago"].map(partner_names)
            diag["ret_diferente_partner"] = diff[[
                c for c in [
                    "move_id", "partner_id", "nombre_ret",
                    "partner_pago", "nombre_pago",
                    "account_code", "monto_ret",
                ] if c in diff.columns
            ]]

    return diag


# ===========================================================================
# FORMATO 1003 — Retenciones que NOS practicaron (en ventas)
# ===========================================================================


def build_formato_1003(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    partners: pd.DataFrame,
    year: int,
) -> pd.DataFrame:
    """
    Retenciones que clientes practicaron a la empresa.

    Cuentas típicas: 1355 (Anticipo de impuestos y contribuciones).
    """
    if moves is None or moves.empty:
        return pd.DataFrame()
    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.year == year]
    if df.empty:
        return pd.DataFrame()
    df = _enrich_moves_with_puc(df, chart)

    # 1355xx = anticipos y retenciones practicadas a la empresa
    ret_mask = df["account_code"].str.startswith("1355")
    rets = df[ret_mask & (df["partner_id"].notna())].copy()
    if rets.empty:
        return pd.DataFrame()
    rets["monto_ret"] = rets["debit"].fillna(0) - rets["credit"].fillna(0)
    grp = _group_partner_with_cuentas(rets, "monto_ret")

    if partners is not None and not partners.empty:
        partner_info = partners.set_index("id")
        out_rows = []
        for _, r in grp.iterrows():
            pid = r["partner_id"]
            try:
                p = partner_info.loc[int(pid)].to_dict() if pid in partner_info.index else {}
            except Exception:  # noqa: BLE001
                p = {}
            t = _row_tercero(p)
            out_rows.append({
                **t,
                "retencion_practicada": round(float(r["monto_ret"]), 0),
            })
        return pd.DataFrame(out_rows)
    return grp


# ===========================================================================
# FORMATO 1004 — Descuentos tributarios
# ===========================================================================


def build_formato_1004(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    partners: pd.DataFrame,
    year: int,
) -> pd.DataFrame:
    """
    Descuentos tributarios: cuentas 1715 (Impuestos diferidos) o 1620
    (Descuentos tributarios). Devuelve por tercero.
    """
    if moves is None or moves.empty:
        return pd.DataFrame()
    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.year == year]
    if df.empty:
        return pd.DataFrame()
    df = _enrich_moves_with_puc(df, chart)

    # Cuentas típicas para descuentos tributarios
    mask = df["account_code"].str.startswith(("1715", "1620"))
    sub = df[mask & (df["partner_id"].notna())].copy()
    if sub.empty:
        return pd.DataFrame()
    sub["monto"] = sub["debit"].fillna(0) - sub["credit"].fillna(0)
    grp = sub.groupby("partner_id", as_index=False)["monto"].sum()
    if grp.empty:
        return pd.DataFrame()
    return _enrich_with_partner(grp, partners, "monto", "descuento_tributario")


# ===========================================================================
# FORMATO 1005 — IVA descontable (compras)
# ===========================================================================


def build_formato_1005(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    partners: pd.DataFrame,
    year: int,
) -> pd.DataFrame:
    """
    IVA descontable por proveedor. Cuenta 2408 = IVA descontable.
    """
    if moves is None or moves.empty:
        return pd.DataFrame()
    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.year == year]
    if df.empty:
        return pd.DataFrame()
    df = _enrich_moves_with_puc(df, chart)

    mask = df["account_code"].str.startswith("2408")
    sub = df[mask & (df["partner_id"].notna())].copy()
    if sub.empty:
        return pd.DataFrame()
    # IVA descontable: el descuento se acredita (credit) cuando es a favor
    sub["iva_descontable"] = sub["debit"].fillna(0) - sub["credit"].fillna(0)
    grp = _group_partner_with_cuentas(sub, "iva_descontable")
    grp = grp[grp["iva_descontable"].abs() >= 1]
    return _enrich_with_partner(grp, partners, "iva_descontable", "iva_descontable")


# ===========================================================================
# FORMATO 1006 — IVA generado (ventas)
# ===========================================================================


def build_formato_1006(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    partners: pd.DataFrame,
    year: int,
) -> pd.DataFrame:
    """IVA generado por cliente. Cuenta 2408 cuando es por ventas."""
    if moves is None or moves.empty:
        return pd.DataFrame()
    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.year == year]
    if df.empty:
        return pd.DataFrame()
    df = _enrich_moves_with_puc(df, chart)

    mask = df["account_code"].str.startswith(("2408", "240805"))
    sub = df[mask & (df["partner_id"].notna())].copy()
    if sub.empty:
        return pd.DataFrame()
    # IVA generado: crédito a 2408 cuando se factura
    sub["iva_generado"] = sub["credit"].fillna(0) - sub["debit"].fillna(0)
    grp = _group_partner_with_cuentas(sub, "iva_generado")
    grp = grp[grp["iva_generado"] > 0]
    return _enrich_with_partner(grp, partners, "iva_generado", "iva_generado")


# ===========================================================================
# FORMATO 1007 — Ingresos recibidos
# ===========================================================================


def build_formato_1007(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    partners: pd.DataFrame,
    year: int,
) -> pd.DataFrame:
    """
    Ingresos por tercero (cliente). Suma los créditos a cuentas grupo 4
    (ingresos) del año fiscal, agrupado por (tercero, concepto).
    """
    if moves is None or moves.empty:
        return pd.DataFrame()
    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.year == year]
    if df.empty:
        return pd.DataFrame()
    df = _enrich_moves_with_puc(df, chart)

    # El 1007 reporta INGRESOS — debe venir de facturas de venta o asientos
    # manuales. Excluir facturas de compra (in_invoice/in_refund) que
    # podrían tocar cuentas 4xxx (ej. descuentos de proveedor a 4170).
    if "move_type" in df.columns:
        TIPOS_VALIDOS_1007 = {
            "out_invoice", "out_refund", "entry", False, None, "",
        }
        df_filt = df[df["move_type"].isin(TIPOS_VALIDOS_1007) | df["move_type"].isna()]
    else:
        df_filt = df

    mask = df_filt["puc_group"] == "4"
    sub = df_filt[mask & (df_filt["partner_id"].notna())].copy()
    if sub.empty:
        return pd.DataFrame()
    sub["ingreso"] = sub["credit"].fillna(0) - sub["debit"].fillna(0)
    sub["concepto"] = sub["account_code"].apply(
        lambda c: _get_concepto(c, CONCEPTOS_1007_POR_PUC)
    )
    grp = _group_partner_with_cuentas(sub, "ingreso", extra_group_cols=["concepto"])
    grp = grp[grp["ingreso"] > 0]
    return _enrich_with_partner(grp, partners, "ingreso", "ingreso_bruto")


# ===========================================================================
# FORMATO 1008 — Cuentas por cobrar al 31 de diciembre
# ===========================================================================


def build_formato_1008(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    partners: pd.DataFrame,
    year: int,
) -> pd.DataFrame:
    """
    Saldo de cuentas por cobrar (subgrupo 13) al 31 dic del año.
    """
    if moves is None or moves.empty:
        return pd.DataFrame()
    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.date <= date(year, 12, 31)]
    if df.empty:
        return pd.DataFrame()
    df = _enrich_moves_with_puc(df, chart)

    # CxC: subgrupo 13 (deudores comerciales) usando PUC robusto + fallback código
    mask = (df["puc_subgroup"] == "13") | df["account_code"].str.startswith((
        "1305", "1310", "1325", "1330",
    ))
    sub = df[mask & (df["partner_id"].notna())].copy()
    if sub.empty:
        return pd.DataFrame()
    sub["saldo"] = sub["debit"].fillna(0) - sub["credit"].fillna(0)
    grp = _group_partner_with_cuentas(sub, "saldo")
    grp = grp[grp["saldo"] > 0]  # solo deudoras
    return _enrich_with_partner(grp, partners, "saldo", "saldo_cuenta_cobrar")


# ===========================================================================
# FORMATO 1009 — Cuentas por pagar al 31 de diciembre
# ===========================================================================


def build_formato_1009(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    partners: pd.DataFrame,
    year: int,
) -> pd.DataFrame:
    """Saldos de CxP (subgrupos 22-23) al 31 dic."""
    if moves is None or moves.empty:
        return pd.DataFrame()
    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.date <= date(year, 12, 31)]
    if df.empty:
        return pd.DataFrame()
    df = _enrich_moves_with_puc(df, chart)

    # CxP: subgrupos 22 (proveedores) y 23 (cuentas por pagar)
    mask = df["puc_subgroup"].isin(["22", "23"]) | df["account_code"].str.startswith((
        "22", "2335", "2380",
    ))
    sub = df[mask & (df["partner_id"].notna())].copy()
    if sub.empty:
        return pd.DataFrame()
    sub["saldo"] = sub["credit"].fillna(0) - sub["debit"].fillna(0)
    sub["concepto"] = sub["account_code"].apply(
        lambda c: _get_concepto(c, CONCEPTOS_1009_POR_PUC)
    )
    grp = _group_partner_with_cuentas(sub, "saldo", extra_group_cols=["concepto"])
    grp = grp[grp["saldo"] > 0]  # solo acreedoras
    return _enrich_with_partner(grp, partners, "saldo", "saldo_cuenta_pagar")


# ===========================================================================
# FORMATO 1010 — Información de socios y accionistas
# ===========================================================================


def build_formato_1010(
    partners: pd.DataFrame,
    year: int,
) -> pd.DataFrame:
    """
    Información de socios/accionistas. Odoo NO marca terceros como socios
    automáticamente. Devuelve una plantilla con todos los partners tipo
    'company' o con campo `is_shareholder` si existe, para que el usuario
    revise y complete.
    """
    if partners is None or partners.empty:
        return pd.DataFrame()
    df = partners.copy()
    # Tomar partners que potencialmente sean socios (heurística simple)
    if "is_shareholder" in df.columns:
        df = df[df["is_shareholder"] == True]  # noqa: E712
    else:
        # Sin campo específico, devolvemos plantilla con TODOS los partners
        # y el usuario debe filtrar manualmente cuáles son socios
        pass
    rows = []
    for _, p in df.iterrows():
        t = _row_tercero(p.to_dict())
        rows.append({
            **t,
            "porcentaje_participacion": 0,  # llenar manualmente
            "valor_patrimonio": 0,
        })
    return pd.DataFrame(rows)


# ===========================================================================
# FORMATO 1011 — Información declaraciones tributarias (consolidado)
# ===========================================================================


def build_formato_1011(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    year: int,
) -> pd.DataFrame:
    """
    Información agregada de la declaración de renta del año.

    Devuelve totales de:
      - Ingresos brutos
      - Costos
      - Gastos operacionales
      - Gastos no operacionales
      - Impuesto al patrimonio (si aplica)
      - Renta exenta
    """
    if moves is None or moves.empty:
        return pd.DataFrame()
    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.year == year]
    if df.empty:
        return pd.DataFrame()
    df = _enrich_moves_with_puc(df, chart)

    def _suma_subgroup(subgroup: str) -> float:
        m = df["puc_subgroup"] == subgroup
        sub = df[m]
        return float(sub["credit"].sum() - sub["debit"].sum())

    def _suma_group(group: str) -> float:
        m = df["puc_group"] == group
        sub = df[m]
        return float(sub["credit"].sum() - sub["debit"].sum())

    rows = [
        ("Concepto", "Valor"),
        ("Ingresos brutos operacionales (41)", _suma_subgroup("41")),
        ("Ingresos no operacionales (42)", _suma_subgroup("42")),
        ("Costo de ventas (6)", -_suma_group("6")),  # invertir signo
        ("Gastos administración (51)", -_suma_subgroup("51")),
        ("Gastos ventas (52)", -_suma_subgroup("52")),
        ("Gastos no operacionales (53)", -_suma_subgroup("53")),
    ]
    return pd.DataFrame(rows[1:], columns=rows[0])


# ===========================================================================
# FORMATO 1012 — Información de saldos (caja, bancos, inversiones)
# ===========================================================================


def build_formato_1012(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    year: int,
) -> pd.DataFrame:
    """
    Saldos al 31 dic de cuentas de:
      - 1105 Caja
      - 1110 Bancos
      - 1120 Cuentas de ahorro
      - 1205 Inversiones
      - 1305 CxC
    """
    if moves is None or moves.empty:
        return pd.DataFrame()
    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.date <= date(year, 12, 31)]
    if df.empty:
        return pd.DataFrame()
    df = _enrich_moves_with_puc(df, chart)

    # Subgrupos 11 (Disponible: caja/bancos) y 12 (Inversiones)
    mask = df["puc_subgroup"].isin(["11", "12"]) | df["account_code"].str.startswith(
        ("11", "12")
    )
    sub = df[mask].copy()
    if sub.empty:
        return pd.DataFrame()
    sub["saldo"] = sub["debit"].fillna(0) - sub["credit"].fillna(0)
    grp = sub.groupby(
        ["account_code", "account_name"], as_index=False,
    )["saldo"].sum()
    return grp.sort_values("account_code").reset_index(drop=True)


# ===========================================================================
# Helper: enriquecer DataFrame con info de tercero
# ===========================================================================


def _enrich_with_partner(
    df: pd.DataFrame,
    partners: pd.DataFrame,
    valor_col: str,
    nombre_final: str,
) -> pd.DataFrame:
    """Agrega columnas de tercero a un DataFrame que tenga partner_id."""
    if partners is None or partners.empty or df.empty:
        return df
    partner_info = partners.set_index("id")
    rows = []
    for _, r in df.iterrows():
        pid = r["partner_id"]
        try:
            p = partner_info.loc[int(pid)].to_dict() if pid in partner_info.index else {}
        except Exception:  # noqa: BLE001
            p = {}
        t = _row_tercero(p)
        row = {**t}
        # Conservar columnas adicionales del df original (ej. concepto, cuentas)
        for c in df.columns:
            if c not in ("partner_id", valor_col):
                row[c] = r[c]
        row[nombre_final] = round(float(r[valor_col]), 0)
        rows.append(row)
    return pd.DataFrame(rows)


def _cuentas_unicas_agg(s: pd.Series) -> str:
    """Helper: agrega códigos/nombres únicos separados por coma.

    Excluye valores vacíos, None y el literal "False" (que aparece
    cuando Odoo devuelve campos nulos vía XML-RPC).
    """
    vals = sorted({
        str(v).strip()
        for v in s.dropna()
        if str(v).strip() and str(v).strip().lower() != "false"
    })
    return ", ".join(vals)


def _group_partner_with_cuentas(
    sub: pd.DataFrame,
    valor_col: str,
    extra_group_cols: list[str] | None = None,
) -> pd.DataFrame:
    """
    Helper: agrupa por partner_id (+ extra_group_cols opcionales),
    suma `valor_col`, y agrega columnas con cuentas únicas para validación.

    Devuelve DataFrame con:
      - partner_id (+ extra_group_cols)
      - valor_col (sumado)
      - cuentas_contables (códigos PUC únicos, separados por coma)
      - nombres_cuentas (nombres de cuentas únicos)
      - n_lineas (cantidad de líneas que componen el grupo)
    """
    if sub is None or sub.empty:
        return pd.DataFrame()
    group_cols = ["partner_id"] + (extra_group_cols or [])
    has_code = "account_code" in sub.columns
    has_name = "account_name" in sub.columns

    agg = {valor_col: "sum"}
    grp = sub.groupby(group_cols, as_index=False).agg(agg)
    # Cuentas
    if has_code:
        cuentas = sub.groupby(group_cols, as_index=False).agg(
            cuentas_contables=("account_code", _cuentas_unicas_agg),
        )
        grp = grp.merge(cuentas, on=group_cols, how="left")
    else:
        grp["cuentas_contables"] = ""
    if has_name:
        nombres = sub.groupby(group_cols, as_index=False).agg(
            nombres_cuentas=("account_name", _cuentas_unicas_agg),
        )
        grp = grp.merge(nombres, on=group_cols, how="left")
    else:
        grp["nombres_cuentas"] = ""
    # n_lineas
    n_lin = sub.groupby(group_cols, as_index=False).size().rename(
        columns={"size": "n_lineas"}
    )
    grp = grp.merge(n_lin, on=group_cols, how="left")
    return grp


# ===========================================================================
# FORMATO 1015 — Pasivos al 31 de diciembre
# ===========================================================================


def build_formato_1015(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    partners: pd.DataFrame,
    year: int,
) -> pd.DataFrame:
    """
    Saldos de pasivos al cierre del año. Incluye:
      - 21xx Obligaciones financieras (préstamos bancarios)
      - 22xx Proveedores
      - 23xx Cuentas por pagar
      - 24xx Impuestos por pagar
      - 25xx Obligaciones laborales

    NOTA: 1015 a veces se usa solo para obligaciones financieras (21xx).
    Aquí devolvemos TODAS las cuentas 21-25 con su tercero.
    """
    if moves is None or moves.empty:
        return pd.DataFrame()
    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.date <= date(year, 12, 31)]
    if df.empty:
        return pd.DataFrame()
    df = _enrich_moves_with_puc(df, chart)

    # Pasivos: puc_group "2" (pasivos) — incluye 21-25
    mask = (df["puc_group"] == "2") | df["account_code"].str.startswith(
        ("21", "22", "23", "24", "25")
    )
    sub = df[mask & (df["partner_id"].notna())].copy()
    if sub.empty:
        return pd.DataFrame()
    sub["saldo_pasivo"] = sub["credit"].fillna(0) - sub["debit"].fillna(0)
    sub["grupo_pasivo"] = sub["account_code"].str[:2]
    grp = _group_partner_with_cuentas(
        sub, "saldo_pasivo", extra_group_cols=["grupo_pasivo"],
    )
    grp = grp[grp["saldo_pasivo"] > 0]
    if grp.empty:
        return pd.DataFrame()
    return _enrich_with_partner(grp, partners, "saldo_pasivo", "saldo_pasivo")


# ===========================================================================
# FORMATO 1056 — Devoluciones, anulaciones, rescisiones
# ===========================================================================


def build_formato_1056(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    partners: pd.DataFrame,
    year: int,
) -> pd.DataFrame:
    """
    Devoluciones en ventas (notas crédito) y compras (notas crédito de
    proveedor) del año, por tercero.

    Identificación: movimientos a cuentas 4175 (Devoluciones en ventas) y
    6225 (Devoluciones en compras), o cualquier débito a 4xxx (NC ventas).
    """
    if moves is None or moves.empty:
        return pd.DataFrame()
    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.year == year]
    if df.empty:
        return pd.DataFrame()
    df = _enrich_moves_with_puc(df, chart)

    # Devoluciones en ventas: cuentas 4175 o débitos a ingresos (puc_group="4")
    mask_dev_v = df["account_code"].str.startswith("4175") | (
        (df["puc_group"] == "4") & (df["debit"] > 0)
    )
    # Devoluciones en compras: cuentas 6225 o créditos a costos (puc_group="6")
    mask_dev_c = df["account_code"].str.startswith("6225") | (
        (df["puc_group"] == "6") & (df["credit"] > 0)
    )

    devoluciones_ventas = df[mask_dev_v & (df["partner_id"].notna())].copy()
    devoluciones_ventas["devolucion_venta"] = devoluciones_ventas["debit"].fillna(0)

    devoluciones_compras = df[mask_dev_c & (df["partner_id"].notna())].copy()
    devoluciones_compras["devolucion_compra"] = devoluciones_compras["credit"].fillna(0)

    rows_v = devoluciones_ventas.groupby("partner_id", as_index=False)[
        "devolucion_venta"
    ].sum()
    rows_c = devoluciones_compras.groupby("partner_id", as_index=False)[
        "devolucion_compra"
    ].sum()
    out = rows_v.merge(rows_c, on="partner_id", how="outer").fillna(0)
    out = out[(out["devolucion_venta"] > 0) | (out["devolucion_compra"] > 0)]
    if out.empty:
        return pd.DataFrame()

    # Enriquecer con info de tercero
    if partners is not None and not partners.empty:
        partner_info = partners.set_index("id")
        out_rows = []
        for _, r in out.iterrows():
            pid = r["partner_id"]
            try:
                p = partner_info.loc[int(pid)].to_dict() if pid in partner_info.index else {}
            except Exception:  # noqa: BLE001
                p = {}
            t = _row_tercero(p)
            out_rows.append({
                **t,
                "devolucion_venta": round(float(r["devolucion_venta"]), 0),
                "devolucion_compra": round(float(r["devolucion_compra"]), 0),
            })
        return pd.DataFrame(out_rows)
    return out


# ===========================================================================
# FORMATO 1647 — Ingresos recibidos para terceros
# ===========================================================================


def build_formato_1647(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    partners: pd.DataFrame,
    year: int,
) -> pd.DataFrame:
    """
    Ingresos que la empresa recibe POR CUENTA de terceros (no son ingresos
    propios). Cuenta típica: 2815 Ingresos recibidos para terceros.
    """
    if moves is None or moves.empty:
        return pd.DataFrame()
    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.year == year]
    if df.empty:
        return pd.DataFrame()
    df = _enrich_moves_with_puc(df, chart)

    mask = df["account_code"].str.startswith("2815")
    sub = df[mask & (df["partner_id"].notna())].copy()
    if sub.empty:
        return pd.DataFrame()
    sub["ingreso_terceros"] = sub["credit"].fillna(0) - sub["debit"].fillna(0)
    grp = _group_partner_with_cuentas(sub, "ingreso_terceros")
    grp = grp[grp["ingreso_terceros"] > 0]
    if grp.empty:
        return pd.DataFrame()
    return _enrich_with_partner(grp, partners, "ingreso_terceros", "ingreso_recibido_terceros")


# ===========================================================================
# FORMATO 2275 — Costos y deducciones (resolución reciente DIAN)
# ===========================================================================


def build_formato_2275(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    partners: pd.DataFrame,
    year: int,
) -> pd.DataFrame:
    """
    Costos y deducciones por tercero — formato más reciente y detallado
    que 1001. Toma cuentas 5xxx, 6xxx, 7xxx con tercero y agrupa por
    (partner, código contable).

    Diferencia con 1001: este reporta el NIVEL DE CUENTA específico
    (account_code), no el concepto DIAN.
    """
    if moves is None or moves.empty:
        return pd.DataFrame()
    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.year == year]
    if df.empty:
        return pd.DataFrame()
    df = _enrich_moves_with_puc(df, chart)

    # Costos y deducciones: puc_group 5 (gastos), 6 (costos), 7 (costos producción)
    mask = df["puc_group"].isin(["5", "6", "7"])
    sub = df[mask & (df["partner_id"].notna())].copy()
    if sub.empty:
        return pd.DataFrame()
    sub["monto"] = sub["debit"].fillna(0) - sub["credit"].fillna(0)
    grp = sub.groupby(
        ["partner_id", "account_code", "account_name"], as_index=False,
    )["monto"].sum()
    grp = grp[grp["monto"].abs() >= 1]
    if grp.empty:
        return pd.DataFrame()
    return _enrich_with_partner(grp, partners, "monto", "valor_costo_deduccion")


# ===========================================================================
# FORMATO 2276 — Pagos laborales (rentas de trabajo)
# ===========================================================================


def build_formato_2276(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    partners: pd.DataFrame,
    year: int,
) -> pd.DataFrame:
    """
    Pagos laborales detallados por empleado. Toma cuentas 5105 (Gastos
    de personal), 7105 (Costos de personal), 2510 (Cesantías), etc.

    NOTA: en Odoo los empleados no siempre están en res.partner. Si el
    cliente usa el módulo HR, los datos están en hr.payslip — pero este
    informe usa el movimiento contable como fuente.
    """
    if moves is None or moves.empty:
        return pd.DataFrame()
    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.year == year]
    if df.empty:
        return pd.DataFrame()
    df = _enrich_moves_with_puc(df, chart)

    # Cuentas laborales: 5105, 5108, 5110-laborales, 7105, 2510-2530
    mask_lab = df["account_code"].str.startswith((
        "5105", "5108", "5106", "7105", "7108",
        "2510", "2515", "2520", "2525", "2530",
    ))
    sub = df[mask_lab & (df["partner_id"].notna())].copy()
    if sub.empty:
        return pd.DataFrame()
    sub["monto_laboral"] = sub["debit"].fillna(0) - sub["credit"].fillna(0)

    # Clasificar tipo: salario, prestaciones, seguridad social
    def _tipo_laboral(code: str) -> str:
        if code.startswith(("5105", "7105")):
            return "Salarios"
        if code.startswith(("5106", "7106", "2510")):
            return "Cesantías e intereses"
        if code.startswith(("5108", "7108", "2515", "2520")):
            return "Prestaciones sociales"
        if code.startswith(("2525", "2530")):
            return "Aportes seguridad social"
        return "Otros laborales"

    sub["tipo_pago"] = sub["account_code"].apply(_tipo_laboral)
    grp = _group_partner_with_cuentas(
        sub, "monto_laboral", extra_group_cols=["tipo_pago"],
    )
    grp = grp[grp["monto_laboral"] > 0]
    if grp.empty:
        return pd.DataFrame()
    return _enrich_with_partner(grp, partners, "monto_laboral", "valor_pagado")


# ===========================================================================
# Generación del Excel multi-hoja con todos los formatos
# ===========================================================================


def build_validacion_por_cuenta(
    moves: pd.DataFrame,
    chart: pd.DataFrame,
    partners: pd.DataFrame,
    year: int,
) -> pd.DataFrame:
    """
    Construye una hoja de validación cruzada: para CADA cuenta contable
    que participa en algún formato DIAN, suma el saldo del año y lo
    desglosa por tercero. Sirve para cuadrar contra el libro mayor.

    Devuelve DataFrame con columnas:
      - account_code, account_name
      - partner_id, partner_name
      - debit_total, credit_total, saldo_neto
      - n_lineas
    """
    if moves is None or moves.empty:
        return pd.DataFrame()
    df = moves.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].dt.year == year]
    if df.empty:
        return pd.DataFrame()
    df = _enrich_moves_with_puc(df, chart)

    sub = df[df["partner_id"].notna()].copy()
    if sub.empty:
        return pd.DataFrame()
    sub["debit"] = sub["debit"].fillna(0)
    sub["credit"] = sub["credit"].fillna(0)
    sub["saldo_neto"] = sub["debit"] - sub["credit"]

    grp = sub.groupby(
        ["account_code", "account_name", "partner_id"],
        as_index=False,
    ).agg(
        debit_total=("debit", "sum"),
        credit_total=("credit", "sum"),
        saldo_neto=("saldo_neto", "sum"),
        n_lineas=("debit", "size"),
    )
    # Enriquecer con nombre del partner
    if partners is not None and not partners.empty:
        partner_names = partners.set_index("id")["name"].to_dict()
        grp["partner_name"] = grp["partner_id"].map(partner_names)
    else:
        grp["partner_name"] = ""
    # Reordenar columnas
    cols = [
        "account_code", "account_name",
        "partner_id", "partner_name",
        "debit_total", "credit_total", "saldo_neto", "n_lineas",
    ]
    return grp[[c for c in cols if c in grp.columns]].sort_values(
        ["account_code", "saldo_neto"], ascending=[True, False],
    ).reset_index(drop=True)


# ===========================================================================
# Configuración de exportación Excel — mapeo de columnas a nombres DIAN
# ===========================================================================
#
# Para cada formato definimos:
#   columnas_dian: orden y nombres DIAN-friendly (lo que va al reporte oficial)
#   columnas_auditoria: columnas extra que añadimos para revisar/validar
#                       (cuentas usadas, # líneas, IDs internos)
#
# Las columnas comunes de "tercero" tienen estos nombres internos:
#   tipo_doc, numero_doc, dv, apellido1, apellido2, nombre1, nombre2,
#   razon_social, direccion, departamento, municipio, pais

COLS_TERCERO_DIAN: list[tuple[str, str]] = [
    ("tipo_doc", "Tipo Documento"),
    ("numero_doc", "Número Identificación"),
    ("dv", "DV"),
    ("apellido1", "Primer Apellido"),
    ("apellido2", "Segundo Apellido"),
    ("nombre1", "Primer Nombre"),
    ("nombre2", "Otros Nombres"),
    ("razon_social", "Razón Social"),
    ("direccion", "Dirección"),
    ("departamento", "Departamento"),
    ("municipio", "Municipio"),
    ("pais", "País"),
]

COLS_AUDITORIA: list[tuple[str, str]] = [
    ("cuentas_contables", "Cuentas Contables (auditoría)"),
    ("nombres_cuentas", "Nombres Cuentas (auditoría)"),
    ("n_lineas", "# Líneas (auditoría)"),
    ("partner_id", "ID Tercero (interno)"),
]

# Mapeo específico por formato: monto + columnas adicionales
SCHEMA_POR_FORMATO: dict[str, list[tuple[str, str, str]]] = {
    # (col_interna, col_dian, tipo: 'money' | 'text' | 'int' | 'pct')
    "1001": [
        ("concepto", "Concepto", "text"),
        ("pago_deducible", "Pago o Abono Deducible", "money"),
        ("pago_no_deducible", "Pago o Abono No Deducible", "money"),
        ("ret_fuente_renta", "Retención en la Fuente Renta", "money"),
        ("ret_iva", "Retención IVA", "money"),
        ("ret_ica", "Retención ICA", "money"),
    ],
    "1003": [
        ("retencion_practicada", "Retención que Practicaron", "money"),
    ],
    "1004": [
        ("descuento_tributario", "Descuento Tributario", "money"),
    ],
    "1005": [
        ("iva_descontable", "IVA Descontable", "money"),
    ],
    "1006": [
        ("iva_generado", "IVA Generado", "money"),
    ],
    "1007": [
        ("concepto", "Concepto", "text"),
        ("ingreso_bruto", "Ingreso Bruto Recibido", "money"),
    ],
    "1008": [
        ("saldo_cuenta_cobrar", "Saldo Cuenta por Cobrar", "money"),
    ],
    "1009": [
        ("concepto", "Concepto", "text"),
        ("saldo_cuenta_pagar", "Saldo Cuenta por Pagar", "money"),
    ],
    "1010": [
        ("porcentaje_participacion", "% Participación", "pct"),
        ("valor_aportes", "Valor Aportes", "money"),
    ],
    "1011": [
        ("concepto", "Concepto", "text"),
        ("valor", "Valor", "money"),
    ],
    "1012": [
        ("concepto", "Concepto", "text"),
        ("saldo", "Saldo a 31 Diciembre", "money"),
    ],
    "1015": [
        ("concepto", "Concepto", "text"),
        ("saldo", "Saldo Pasivo", "money"),
    ],
    "1056": [
        ("concepto", "Concepto", "text"),
        ("valor", "Valor Devolución/Anulación", "money"),
    ],
    "1647": [
        ("concepto", "Concepto", "text"),
        ("ingreso_terceros", "Ingreso para Terceros", "money"),
    ],
    "2275": [
        ("concepto", "Concepto", "text"),
        ("valor", "Costo/Deducción", "money"),
    ],
    "2276": [
        ("concepto", "Concepto", "text"),
        ("pago_laboral", "Pago Laboral", "money"),
    ],
}


def _formato_code_from_key(key: str) -> str:
    """Extrae el código del formato ('1001') del título completo."""
    return key.split(" ")[0].strip()


def _reordenar_columnas_dian(
    df: pd.DataFrame,
    formato_code: str,
) -> tuple[pd.DataFrame, dict[str, str]]:
    """
    Reordena las columnas del DataFrame según el orden DIAN oficial
    y retorna (df_ordenado, mapeo_nombres) donde mapeo_nombres lleva
    nombres internos → nombres DIAN-friendly.
    """
    if df is None or df.empty:
        return df, {}

    schema = SCHEMA_POR_FORMATO.get(formato_code, [])

    # Orden: tercero → schema-específico → auditoría
    orden_cols: list[str] = []
    nombres: dict[str, str] = {}

    for internal, dian in COLS_TERCERO_DIAN:
        if internal in df.columns:
            orden_cols.append(internal)
            nombres[internal] = dian

    for internal, dian, _tipo in schema:
        if internal in df.columns:
            orden_cols.append(internal)
            nombres[internal] = dian

    # Cualquier columna que no esté mapeada explícitamente
    cols_no_mapeadas = [
        c for c in df.columns
        if c not in orden_cols
        and c not in [x[0] for x in COLS_AUDITORIA]
    ]
    orden_cols.extend(cols_no_mapeadas)
    for c in cols_no_mapeadas:
        nombres[c] = c

    # Auditoría al final
    for internal, dian in COLS_AUDITORIA:
        if internal in df.columns:
            orden_cols.append(internal)
            nombres[internal] = dian

    df_out = df[orden_cols].copy()
    return df_out, nombres


def _aplicar_formato_excel(worksheet, df: pd.DataFrame,
                            formato_code: str, nombres: dict[str, str]) -> None:
    """Aplica formato a una hoja: header bold, freeze, números, anchos."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    # Header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, _col_name in enumerate(df.columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Freeze header
    worksheet.freeze_panes = "A2"

    # Formato numérico para columnas money/int/pct
    schema = SCHEMA_POR_FORMATO.get(formato_code, [])
    tipo_por_col_dian: dict[str, str] = {dian: tipo for _, dian, tipo in schema}
    # Columnas de auditoría que son numéricas
    tipo_por_col_dian["# Líneas (auditoría)"] = "int"
    tipo_por_col_dian["ID Tercero (interno)"] = "int"
    tipo_por_col_dian["DV"] = "text"
    tipo_por_col_dian["Tipo Documento"] = "text"

    formato_money = '#,##0;[Red]-#,##0'
    formato_int = '#,##0'
    formato_pct = '0.00%'

    for col_idx, col_name in enumerate(df.columns, start=1):
        tipo = tipo_por_col_dian.get(col_name, "text")
        col_letter = get_column_letter(col_idx)
        if tipo == "money":
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=col_idx).number_format = formato_money
        elif tipo == "int":
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=col_idx).number_format = formato_int
        elif tipo == "pct":
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=col_idx).number_format = formato_pct

        # Auto-width aproximado (max length de los valores en la columna)
        try:
            max_len = max(
                [len(str(col_name))] +
                [len(str(v)) for v in df.iloc[:, col_idx - 1].fillna("").head(200)]
            )
            ancho = min(max(max_len + 2, 12), 50)
            worksheet.column_dimensions[col_letter].width = ancho
        except Exception:  # noqa: BLE001
            worksheet.column_dimensions[col_letter].width = 18

    # Altura de fila del header
    worksheet.row_dimensions[1].height = 32


def generar_excel_medios_magneticos(
    formatos: dict[str, pd.DataFrame],
    output_path: str,
    year: int,
    validacion: pd.DataFrame | None = None,
) -> None:
    """
    Escribe todos los formatos en un Excel multi-hoja con formato profesional:
      - Nombres de columnas DIAN-friendly (en español, claros)
      - Orden de columnas: identificación tercero → datos del formato → auditoría
      - Formato de moneda en columnas de pagos/retenciones/saldos
      - Header en negrita con fondo azul oscuro, fila congelada
      - Auto-ancho de columnas
      - Hoja 00_Resumen con totales por formato
      - Hoja ZZ_Validación con cruce cuenta × tercero (si se pasa)

    Args:
        formatos: dict {nombre_formato: DataFrame}, e.g.
                  {"1001 — Pagos y retenciones practicadas": df1001, ...}
        output_path: ruta del archivo de salida (o BytesIO)
        year: año fiscal reportado
        validacion: DataFrame opcional con cruce cuenta × tercero
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ── Hoja 00: Resumen con totales por formato ──
        resumen_rows = []
        for nombre, df in formatos.items():
            formato_code = _formato_code_from_key(nombre)
            schema = SCHEMA_POR_FORMATO.get(formato_code, [])
            # Total del primer campo monetario del formato
            total_principal = 0.0
            campo_total = ""
            if df is not None and not df.empty:
                for internal, dian, tipo in schema:
                    if tipo == "money" and internal in df.columns:
                        try:
                            total_principal = float(df[internal].sum())
                            campo_total = dian
                            break
                        except Exception:  # noqa: BLE001
                            pass
            resumen_rows.append({
                "Formato": nombre,
                "Año": year,
                "Filas": len(df) if df is not None else 0,
                "Campo Principal": campo_total,
                "Total Principal": total_principal,
                "Estado": ("OK" if (df is not None and not df.empty) else "VACÍO"),
            })
        resumen_df = pd.DataFrame(resumen_rows)
        resumen_df.to_excel(writer, sheet_name="00_Resumen", index=False)

        # Aplicar formato a resumen
        ws_resumen = writer.sheets["00_Resumen"]
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(
                start_color="1ABC9C", end_color="1ABC9C", fill_type="solid",
            )
            for col_idx, _ in enumerate(resumen_df.columns, start=1):
                cell = ws_resumen.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            # Formato monetario en columna "Total Principal"
            total_idx = list(resumen_df.columns).index("Total Principal") + 1
            for row in range(2, ws_resumen.max_row + 1):
                ws_resumen.cell(row=row, column=total_idx).number_format = '#,##0;[Red]-#,##0'
            # Anchos
            anchos = {"Formato": 50, "Año": 10, "Filas": 10,
                      "Campo Principal": 32, "Total Principal": 20, "Estado": 12}
            for col_idx, col_name in enumerate(resumen_df.columns, start=1):
                ws_resumen.column_dimensions[get_column_letter(col_idx)].width = (
                    anchos.get(col_name, 15)
                )
            ws_resumen.freeze_panes = "A2"
            ws_resumen.row_dimensions[1].height = 28
        except Exception:  # noqa: BLE001
            pass

        # ── Una hoja por formato ──
        for nombre, df in formatos.items():
            formato_code = _formato_code_from_key(nombre)
            sheet_name = f"{formato_code}"[:31]  # Excel limit

            if df is None or df.empty:
                pd.DataFrame({"Mensaje": [
                    f"Sin datos para el formato {formato_code} en el año {year}.",
                    "Puede ser normal si la empresa no tiene operaciones aplicables.",
                ]}).to_excel(writer, sheet_name=sheet_name, index=False)
                continue

            # Reordenar y renombrar columnas
            df_out, nombres = _reordenar_columnas_dian(df, formato_code)
            df_out = df_out.rename(columns=nombres)
            df_out.to_excel(writer, sheet_name=sheet_name, index=False)

            # Aplicar formato
            _aplicar_formato_excel(
                writer.sheets[sheet_name], df_out, formato_code, nombres,
            )

        # ── Hoja de validación al final ──
        if validacion is not None and not validacion.empty:
            validacion.to_excel(
                writer, sheet_name="ZZ_Validacion_cuenta", index=False,
            )
            try:
                ws_val = writer.sheets["ZZ_Validacion_cuenta"]
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter

                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(
                    start_color="E67E22", end_color="E67E22", fill_type="solid",
                )
                for col_idx, _ in enumerate(validacion.columns, start=1):
                    cell = ws_val.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                ws_val.freeze_panes = "A2"
                ws_val.row_dimensions[1].height = 28
                # Anchos por defecto
                for col_idx in range(1, len(validacion.columns) + 1):
                    ws_val.column_dimensions[get_column_letter(col_idx)].width = 22
            except Exception:  # noqa: BLE001
                pass
